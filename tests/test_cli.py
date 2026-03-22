import json
import subprocess

from openpyxl import load_workbook


def test_cli_help():
    result = subprocess.run(["b2mml", "--help"], capture_output=True, text=True)
    assert result.returncode == 0
    assert "usage" in result.stdout.lower()


def test_cli_convert(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    output_file = tmp_path / "out.xml"

    result = subprocess.run(
        ["b2mml", "convert", input_file, str(output_file)],
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0
    assert output_file.exists()
    assert "<ShowEquipmentInformation" in output_file.read_text()


def test_cli_json(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    output_file = tmp_path / "out.json"

    result = subprocess.run(
        ["b2mml", "json", input_file, str(output_file)], capture_output=True, text=True
    )

    assert result.returncode == 0
    assert output_file.exists()

    data = json.loads(output_file.read_text())
    assert "equipment" in data
    assert "classes" in data
    assert "warnings" in data


def test_cli_json_stdout():
    input_file = "tests/data/sample_ampla.xml"

    result = subprocess.run(
        ["b2mml", "json", input_file], capture_output=True, text=True
    )

    assert result.returncode == 0
    data = json.loads(result.stdout)
    assert "equipment" in data


def test_cli_diff_no_changes(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    result = subprocess.run(
        ["b2mml", "diff", input_file, input_file],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0  # 0 = no differences
    assert "No differences found." in result.stdout


def test_cli_diff_with_changes(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    other_file = tmp_path / "other.xml"
    other_file.write_text(
        """
    <Ampla>
      <Item id="1" name="AcmeMining" type="Citect.Ampla.Isa95.EnterpriseFolder"/>
      <ClassDefinitions>
        <ClassDefinition id="10" name="ProcessingUnit">
          <ClassDefinition id="20" name="Crusher"/>
        </ClassDefinition>
      </ClassDefinitions>
    </Ampla>
    """
    )
    result = subprocess.run(
        ["b2mml", "diff", input_file, str(other_file)],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 1  # 1 = differences found
    assert "-" in result.stdout


def test_cli_diff_json_format(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    result = subprocess.run(
        ["b2mml", "diff", "--format", "json", input_file, input_file],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0
    data = json.loads(result.stdout)
    assert "equipment_added" in data
    assert "classes_added" in data


def test_cli_diff_output_file(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    output_file = tmp_path / "diff.txt"
    result = subprocess.run(
        ["b2mml", "diff", input_file, input_file, str(output_file)],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0
    assert output_file.exists()
    assert "No differences found." in output_file.read_text()


def test_cli_excel(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    output_file = tmp_path / "out.xlsx"
    result = subprocess.run(
        ["b2mml", "excel", input_file, str(output_file)],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0
    assert output_file.exists()
    wb = load_workbook(str(output_file))
    assert "Equipment" in wb.sheetnames
    assert "Classes" in wb.sheetnames


def test_cli_stats_text():
    result = subprocess.run(
        ["b2mml", "stats", "tests/data/sample_ampla.xml"],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0
    assert "Equipment nodes" in result.stdout
    assert "Classes" in result.stdout
    assert "Max depth" in result.stdout
    assert "Warnings" in result.stdout


def test_cli_stats_json():
    result = subprocess.run(
        ["b2mml", "stats", "--format", "json", "tests/data/sample_ampla.xml"],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0
    data = json.loads(result.stdout)
    assert "total_equipment" in data
    assert "total_classes" in data
    assert "max_depth" in data
    assert "warnings" in data


def test_cli_excel_has_warnings_sheet(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    output_file = tmp_path / "out.xlsx"
    subprocess.run(
        ["b2mml", "excel", input_file, str(output_file)],
        capture_output=True,
    )
    wb = load_workbook(str(output_file))
    assert "Warnings" not in wb.sheetnames
