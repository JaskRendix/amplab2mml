import csv
from io import BytesIO, StringIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api import app

client = TestClient(app)


def test_health():
    response = client.get("/health")
    assert response.status_code == 200


def test_convert_json():
    xml = open("tests/data/sample_ampla.xml").read()

    response = client.post(
        "/convert/json", files={"file": ("sample.xml", xml, "application/xml")}
    )

    assert response.status_code == 200
    data = response.json()
    assert "equipment" in data
    assert "classes" in data


def test_convert_xml():
    xml = open("tests/data/sample_ampla.xml").read()

    response = client.post(
        "/convert/xml", files={"file": ("sample.xml", xml, "application/xml")}
    )

    assert response.status_code == 200
    assert "<ShowEquipmentInformation" in response.text


def test_invalid_xml():
    response = client.post(
        "/convert/json", files={"file": ("bad.xml", "<not-xml>", "application/xml")}
    )
    assert response.status_code in (400, 422)


def test_missing_file():
    response = client.post("/convert/json")
    assert response.status_code in (400, 422)


def test_large_file():
    xml = "<Ampla>" + (" " * 500000) + "</Ampla>"
    response = client.post(
        "/convert/json", files={"file": ("large.xml", xml, "application/xml")}
    )
    assert response.status_code == 200


def test_convert_excel():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/excel", files={"file": ("sample.xml", xml, "application/xml")}
    )
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    wb = load_workbook(BytesIO(response.content))
    assert "Equipment" in wb.sheetnames
    assert "Classes" in wb.sheetnames


def test_convert_excel_invalid_xml():
    response = client.post(
        "/convert/excel", files={"file": ("bad.xml", "<not-xml>", "application/xml")}
    )
    assert response.status_code in (400, 422)


def test_convert_csv_equipment():
    import csv
    from io import StringIO

    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/csv/equipment",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    rows = list(csv.DictReader(StringIO(response.text)))
    assert len(rows) > 0
    assert "full_name" in rows[0]
    assert "level" in rows[0]


def test_convert_csv_classes():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/csv/classes",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    rows = list(csv.DictReader(StringIO(response.text)))
    assert len(rows) > 0
    assert "name" in rows[0]
    assert "parent" in rows[0]


def test_stats():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/stats",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["total_equipment"] > 0
    assert data["total_classes"] > 0
    assert data["max_depth"] > 0
    assert "warnings" in data


def test_stats_invalid_xml():
    response = client.post(
        "/stats",
        files={"file": ("bad.xml", "<not-xml>", "application/xml")},
    )
    assert response.status_code in (400, 422)


def test_convert_csv_invalid_xml():
    response = client.post(
        "/convert/csv/equipment",
        files={"file": ("bad.xml", "<not-xml>", "application/xml")},
    )
    assert response.status_code in (400, 422)


def test_convert_html():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/html",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert response.status_code == 200
    assert "text/html" in response.headers["content-type"]
    assert "<!DOCTYPE html>" in response.text
    assert "AcmeMining" in response.text


def test_convert_html_invalid_xml():
    response = client.post(
        "/convert/html",
        files={"file": ("bad.xml", "<not-xml>", "application/xml")},
    )
    assert response.status_code in (400, 422)


def test_diff_json_no_changes():
    xml = open("tests/data/sample_ampla.xml").read()

    response = client.post(
        "/diff/json",
        files={
            "file_a": ("a.xml", xml, "application/xml"),
            "file_b": ("b.xml", xml, "application/xml"),
        },
    )

    assert response.status_code == 200
    data = response.json()

    assert data["equipment_added"] == []
    assert data["equipment_removed"] == []
    assert data["equipment_level_changed"] == []
    assert data["equipment_properties_changed"] == []
    assert data["classes_added"] == []
    assert data["classes_removed"] == []
    assert data["class_properties_changed"] == []


def test_diff_json_with_changes():
    xml_a = open("tests/data/sample_ampla.xml").read()
    xml_b = xml_a.replace("AcmeMining", "AcmeMiningModified")

    response = client.post(
        "/diff/json",
        files={
            "file_a": ("a.xml", xml_a, "application/xml"),
            "file_b": ("b.xml", xml_b, "application/xml"),
        },
    )

    assert response.status_code == 200
    data = response.json()

    # At least one category must contain changes
    assert any(
        len(data[key]) > 0
        for key in [
            "equipment_added",
            "equipment_removed",
            "equipment_level_changed",
            "equipment_properties_changed",
            "classes_added",
            "classes_removed",
            "class_properties_changed",
        ]
    )


def test_diff_text_no_changes():
    xml = open("tests/data/sample_ampla.xml").read()

    response = client.post(
        "/diff/text",
        files={
            "file_a": ("a.xml", xml, "application/xml"),
            "file_b": ("b.xml", xml, "application/xml"),
        },
    )

    assert response.status_code == 200
    assert "no differences" in response.text.lower() or response.text.strip() == ""


def test_diff_text_invalid_xml():
    response = client.post(
        "/diff/text",
        files={
            "file_a": ("a.xml", "<bad>", "application/xml"),
            "file_b": ("b.xml", "<bad>", "application/xml"),
        },
    )
    assert response.status_code in (400, 422)


def test_request_id_passthrough():
    xml = open("tests/data/sample_ampla.xml").read()

    response = client.post(
        "/convert/json",
        headers={"X-Request-ID": "test-123"},
        files={"file": ("sample.xml", xml, "application/xml")},
    )

    assert response.status_code == 200
    assert response.headers["X-Request-ID"] == "test-123"


def test_request_id_generated():
    xml = open("tests/data/sample_ampla.xml").read()

    response = client.post(
        "/convert/json",
        files={"file": ("sample.xml", xml, "application/xml")},
    )

    assert response.status_code == 200
    assert "X-Request-ID" in response.headers
    assert len(response.headers["X-Request-ID"]) > 0


def test_excel_filename():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/excel",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert (
        "attachment; filename=equipment.xlsx" in response.headers["content-disposition"]
    )


def test_csv_equipment_filename():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/csv/equipment",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert (
        "attachment; filename=equipment.csv" in response.headers["content-disposition"]
    )


def test_csv_classes_filename():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/csv/classes",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert "attachment; filename=classes.csv" in response.headers["content-disposition"]


def test_excel_columns():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/excel",
        files={"file": ("sample.xml", xml, "application/xml")},
    )

    wb = load_workbook(BytesIO(response.content))
    ws = wb["Equipment"]

    headers = [cell.value for cell in ws[1]]

    assert "full_name" in headers
    assert "level" in headers
    assert "class_ids" in headers


def test_html_structure():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/html",
        files={"file": ("sample.xml", xml, "application/xml")},
    )

    text = response.text
    assert "<html" in text.lower()
    assert "<body" in text.lower()
    assert "Equipment" in text  # section header


def test_json_includes_warnings():
    xml = open("tests/data/sample_ampla.xml").read()

    response = client.post(
        "/convert/json",
        files={"file": ("sample.xml", xml, "application/xml")},
    )

    assert response.status_code == 200
    data = response.json()
    assert "warnings" in data
    assert isinstance(data["warnings"], list)


def test_empty_file():
    response = client.post(
        "/convert/json",
        files={"file": ("empty.xml", "", "application/xml")},
    )
    assert response.status_code in (400, 422)


def test_non_xml_content():
    response = client.post(
        "/convert/json",
        files={"file": ("weird.txt", "not xml at all", "text/plain")},
    )
    assert response.status_code in (400, 422)


def test_excel_is_binary():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/excel",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert isinstance(response.content, bytes)


def test_all_endpoints_return_request_id():
    xml = open("tests/data/sample_ampla.xml").read()
    endpoints = [
        "/convert/json",
        "/convert/xml",
        "/convert/excel",
        "/convert/csv/equipment",
        "/convert/csv/classes",
        "/convert/html",
        "/stats",
    ]
    for ep in endpoints:
        response = client.post(ep, files={"file": ("x.xml", xml, "application/xml")})
        assert "X-Request-ID" in response.headers


def test_html_contains_equipment_section():
    xml = open("tests/data/sample_ampla.xml").read()
    response = client.post(
        "/convert/html",
        files={"file": ("sample.xml", xml, "application/xml")},
    )

    assert response.status_code == 200
    text = response.text

    # The exact <h2> tag is not guaranteed; check for the section label
    assert "Equipment" in text


def test_diff_json_detects_name_change_only():
    xml_a = open("tests/data/sample_ampla.xml").read()
    xml_b = xml_a.replace("AcmeMining", "X")

    response = client.post(
        "/diff/json",
        files={
            "file_a": ("a.xml", xml_a, "application/xml"),
            "file_b": ("b.xml", xml_b, "application/xml"),
        },
    )

    assert response.status_code == 200
    data = response.json()

    # Name change is not a property diff
    assert data["equipment_properties_changed"] == []

    # Entire subtree is removed and re-added
    assert len(data["equipment_removed"]) > 1
    assert len(data["equipment_added"]) > 1

    # No class-level changes
    assert data["classes_added"] == []
    assert data["classes_removed"] == []
    assert data["class_properties_changed"] == []
