import subprocess
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from lxml import etree
from openpyxl import load_workbook

from app.api import app
from app.excel_export import export_to_excel

XML = """
<Ampla>
  <Item id="1" name="Mine" type="Citect.Ampla.Isa95.EnterpriseFolder">
    <Item id="2" name="Plant" type="Citect.Ampla.Isa95.SiteFolder">
      <ItemClassAssociation classDefinitionId="20"/>
      <Property name="Class.DriveType">Electric</Property>
    </Item>
  </Item>
  <ClassDefinitions>
    <ClassDefinition id="10" name="Base">
      <ClassDefinition id="20" name="Crusher">
        <PropertyDefinition name="DriveType" type="System.String">Unknown</PropertyDefinition>
        <PropertyDefinition name="Manufacturer" type="System.String">ACME</PropertyDefinition>
      </ClassDefinition>
    </ClassDefinition>
  </ClassDefinitions>
</Ampla>
"""


@pytest.fixture
def workbook(make_model):
    model = make_model(XML)
    data = export_to_excel(model)
    return load_workbook(BytesIO(data))


def test_sheets_exist(workbook):
    assert "Equipment" in workbook.sheetnames
    assert "Classes" in workbook.sheetnames


def test_equipment_headers(workbook):
    ws = workbook["Equipment"]
    headers = [cell.value for cell in ws[1]]
    assert headers[:3] == ["full_name", "level", "class_ids"]
    assert "DriveType" in headers
    assert "Manufacturer" in headers


def test_equipment_rows(workbook):
    ws = workbook["Equipment"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    full_names = [r[0] for r in rows]
    assert "Mine" in full_names
    assert "Mine.Plant" in full_names


def test_equipment_property_value(workbook):
    ws = workbook["Equipment"]
    headers = [cell.value for cell in ws[1]]
    drive_col = headers.index("DriveType")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    plant_row = next(r for r in rows if r[0] == "Mine.Plant")
    assert plant_row[drive_col] == "Electric"


def test_equipment_class_ids(workbook):
    ws = workbook["Equipment"]
    headers = [cell.value for cell in ws[1]]
    class_col = headers.index("class_ids")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    plant_row = next(r for r in rows if r[0] == "Mine.Plant")
    assert "Crusher" in plant_row[class_col]


def test_classes_headers(workbook):
    ws = workbook["Classes"]
    headers = [cell.value for cell in ws[1]]
    assert headers[:2] == ["name", "parent"]
    assert "DriveType" in headers


def test_classes_rows(workbook):
    ws = workbook["Classes"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    names = [r[0] for r in rows]
    assert "Crusher" in names


def test_classes_property_value(workbook):
    ws = workbook["Classes"]
    headers = [cell.value for cell in ws[1]]
    mfr_col = headers.index("Manufacturer")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    crusher_row = next(r for r in rows if r[0] == "Crusher")
    assert crusher_row[mfr_col] == "ACME"


def test_returns_bytes(make_model):
    model = make_model(XML)
    result = export_to_excel(model)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_api_export_excel():
    client = TestClient(app)
    xml = open("tests/data/sample_ampla.xml", "rb").read()
    response = client.post(
        "/convert/excel",
        files={"file": ("sample.xml", xml, "application/xml")},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(BytesIO(response.content))
    assert "Equipment" in wb.sheetnames
    assert "Classes" in wb.sheetnames


def test_cli_excel(tmp_path):
    input_file = "tests/data/sample_ampla.xml"
    output_file = tmp_path / "out.xlsx"
    result = subprocess.run(
        ["b2mml", "excel", input_file, str(output_file)],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0
    assert output_file.exists()
    wb = load_workbook(str(output_file))
    assert "Equipment" in wb.sheetnames
