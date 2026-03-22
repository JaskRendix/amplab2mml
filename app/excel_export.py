from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def export_to_excel(model: dict) -> bytes:
    wb = Workbook()

    _write_equipment_sheet(wb, model)
    _write_classes_sheet(wb, model)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _flatten_equipment(equipment_list) -> list:
    flat = []
    stack = list(equipment_list)
    while stack:
        eq = stack.pop(0)
        flat.append(eq)
        stack = list(eq.children) + stack
    return flat


def _write_equipment_sheet(wb: Workbook, model: dict):
    ws = wb.active
    ws.title = "Equipment"

    flat = _flatten_equipment(model["equipment"])

    # collect all property names across all equipment for columns
    prop_names = []
    seen = set()
    for eq in flat:
        for p in eq.properties:
            if p.name not in seen:
                prop_names.append(p.name)
                seen.add(p.name)

    headers = ["full_name", "level", "class_ids"] + prop_names
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for eq in flat:
        prop_map = {p.name: p.value or "" for p in eq.properties}
        row = [
            eq.full_name or "",
            eq.level or "",
            ", ".join(eq.class_ids),
        ] + [prop_map.get(n, "") for n in prop_names]
        ws.append(row)


def _write_classes_sheet(wb: Workbook, model: dict):
    ws = wb.create_sheet(title="Classes")

    classes = model["classes"]

    prop_names = []
    seen = set()
    for cls in classes:
        for p in cls.properties:
            if p.name not in seen:
                prop_names.append(p.name)
                seen.add(p.name)

    headers = ["name", "parent"] + prop_names
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cls in classes:
        prop_map = {p.name: p.value or "" for p in cls.properties}
        row = [
            cls.name,
            cls.parent or "",
        ] + [prop_map.get(n, "") for n in prop_names]
        ws.append(row)
